from fastapi import FastAPI, File, UploadFile, Form, HTTPException
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import pandas as pd
import uvicorn
import tempfile
import shutil
import os
import re
import logging
import xlrd
import subprocess
import mimetypes
import os
import uuid

app = FastAPI()

# Allow CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"]
)

def clean_column_name(col):
    return re.sub(r'[^\w]', '', str(col).lower().strip())

def validate_date(date_str):
    try:
        date_obj = datetime.strptime(date_str, "%d/%m/%y")
        month_year = date_obj.strftime("%B %Y").upper()
        return date_obj, month_year
    except ValueError:
        raise ValueError("Invalid date format. Please use (dd/mm/yy).")

def get_gl_account(description, department_code):
    if str(int(department_code)) in ['3003', '3006']:
        account_map = {
            'TOTAL BASIC SALARY': '640100',
            'Retroactive Appraisal/Arrears': '640100',
            'MONTHLY FOOD': '640140',
            'MONTHLY TRANSP': '640142',
            'MONTHLY HOUSING': '640143',
            'MONTHLY OTHER ALL': '640141',
            'Educatin All': '701200',
            'MONTHLY OVER TIME': '640120'
        }
    else:
        account_map = {
            'TOTAL BASIC SALARY': '701100',
            'Retroactive Appraisal/Arrears': '701100',
            'MONTHLY FOOD': '701210',
            'MONTHLY TRANSP': '701220',
            'MONTHLY HOUSING': '701230',
            'MONTHLY OTHER ALL': '701200',
            'Educatin All': '701200',
            'MONTHLY OVER TIME': '701150'
        }

    base_description = re.sub(r'\s+\w+\s+\d{4}$', '', description).strip()
    return account_map.get(base_description, '')

def convert_xls_to_xlsx(input_path: str) -> str:
    """Converts XLS to XLSX and returns path of converted file"""
    # Create temp output file
    output_path = os.path.join(tempfile.gettempdir(), f"converted_{uuid.uuid4()}.xlsx")
    
    # Run Java converter
    result = subprocess.run(
        [
            "java", 
            "-jar", 
            "xls-xlsx-converter/target/xls-xlsx-converter-1.0-jar-with-dependencies.jar",
            input_path,
            output_path
        ],
        capture_output=True,
        text=True
    )
    
    # Check if conversion succeeded
    if result.returncode != 0:
        raise RuntimeError(f"Conversion failed: {result.stderr}")
    
    if not os.path.exists(output_path):
        raise FileNotFoundError(f"Converted file not found at {output_path}")
    
    return output_path

@app.get("/check-environment")
async def check_env():
    """Verify Java and converter are properly installed"""
    checks = {
        "java": subprocess.run(["which", "java"], capture_output=True),
        "java_version": subprocess.run(["java", "-version"], capture_output=True),
        "converter_jar": os.path.exists("xls-xlsx-converter/target/xls-xlsx-converter-1.0-jar-with-dependencies.jar")
    }
    
    return {
        "java_installed": checks["java"].returncode == 0,
        "java_version": checks["java_version"].stderr.decode().split('\n')[0],
        "converter_jar_exists": checks["converter_jar"],
        "java_home": os.getenv("JAVA_HOME"),
        "path": os.getenv("PATH")
    }

@app.get("/")
def health_check():
    return {"status": "API is up"}

@app.post("/upload/")
async def upload_excel(
    file: UploadFile = File(...),
    sheet_name: str = Form(...),
    posting_date: str = Form(...),
    journal_code: str = Form(...)
):  
    # temp_input = None
    # converted_file = None

    try:
        # Step 1: Validate extension
        ext = os.path.splitext(file.filename)[1].lower()
        if ext not in [".xls", ".xlsx"]:
            raise HTTPException(status_code=400, detail="Only .xls and .xlsx files are supported.")

        # Save uploaded file temporarily
        temp_input = tempfile.NamedTemporaryFile(delete=False, suffix=ext)
        with temp_input as f:
            shutil.copyfileobj(file.file, f)

       # Step 3: If XLS, convert via Java and use converted path
        if ext == ".xls":
            #print(f"Inside xls section - file name {converted_file.name}")
            
            converted_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            java_command = [
                "java","-Xmx1024m" ,"-jar", "xls-xlsx-converter/target/xls-xlsx-converter-1.0-jar-with-dependencies.jar",
                temp_input.name,
                converted_file.name
            ]
            result = subprocess.run(java_command, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            print("Java STDOUT:", result.stdout.decode())
            print("Java STDERR:", result.stderr.decode())
            print("Java return code:", result.returncode)
            print("Expected .xlsx file path:", converted_file.name)

            if result.returncode != 0:
                raise HTTPException(status_code=500, detail=f"Java conversion failed: {result.stderr.decode()}")

            # Sanity check the file MIME type
            mime_type, _ = mimetypes.guess_type(converted_file.name)
            print("MIME type of converted file:", mime_type)
            
            if not mime_type or not mime_type.endswith("spreadsheetml.sheet"):
                raise HTTPException(status_code=500, detail="Converted file is not a valid .xlsx")

            # Confirm file exists and is not zero-size
            if not os.path.exists(converted_file.name) or os.path.getsize(converted_file.name) == 0:
                raise HTTPException(status_code=500, detail="Converted file is empty or missing")

            excel_path = converted_file.name
            print(converted_file.name)
        else:
            print("Inside xlsx section")
            excel_path = temp_input.name
            print(excel_path)

        # Step 4: Validate date
        date_obj, month_year = validate_date(posting_date)
        month_abbr = date_obj.strftime("%b")
        year = date_obj.strftime("%Y")
        new_sheet_name = f"JV {month_abbr} {year}"

         # Step 5: Load workbook and validate sheet
        excel_file = pd.ExcelFile(excel_path, engine="openpyxl")
        sheet_map = {sheet.strip(): sheet for sheet in excel_file.sheet_names}
        actual_sheet_name = sheet_map.get(sheet_name.strip())
        if actual_sheet_name is None:
            raise HTTPException(status_code=400, detail=f"Sheet '{sheet_name}' not found.")
        
        # Step 6: Load dataframe and clean columns
        df = pd.read_excel(excel_path, sheet_name=actual_sheet_name)
        df.columns = [clean_column_name(col) for col in df.columns]

        group_col = 'departmentcode'
        if group_col not in df.columns:
            raise HTTPException(status_code=400, detail="Missing 'Department Code' column.")

        agg_columns = {
            'totalbasicsalary': 'TOTAL BASIC SALARY',
            'retroactiveappraisalarrears': 'Retroactive Appraisal/Arrears',
            'monthlyfood': 'MONTHLY FOOD',
            'monthlytransp': 'MONTHLY TRANSP',
            'monthlyhousing': 'MONTHLY HOUSING',
            'monthlyotherall': 'MONTHLY OTHER ALL',
            'educatinall': 'Educatin All',
            'monthlyovertime': 'MONTHLY OVER TIME'
        }

        available_agg = {k: v for k, v in agg_columns.items() if k in df.columns}
        if not available_agg:
            raise HTTPException(status_code=400, detail="No required aggregation columns found.")

        for col in available_agg.keys():
            df[col] = pd.to_numeric(df[col], errors='coerce')

        agg_df = df.groupby(group_col)[list(available_agg.keys())].sum().reset_index()
        melted_df = agg_df.melt(
            id_vars=[group_col],
            value_vars=list(available_agg.keys()),
            var_name='Description',
            value_name='Amount'
        )
        melted_df['Description'] = melted_df['Description'].map(available_agg) + " " + month_year
        melted_df['Account Number'] = melted_df.apply(
            lambda row: get_gl_account(row['Description'], row[group_col]), axis=1
        )
        melted_df['Posting Date'] = posting_date
        melted_df['Journal Code'] = journal_code
        melted_df['G/L Account'] = "G/L Account"
        melted_df['Department Code'] = melted_df[group_col]

        melted_df.sort_values(by='Department Code', inplace=True)

        final_cols = [
            'Posting Date', 'Journal Code', 'G/L Account',
            'Department Code', 'Account Number', 'Description', 'Amount'
        ]

        # Use openpyxl to preserve formatting
        wb = load_workbook(excel_path)

        # Hide JSR sheet if it exists
        if "JSR" in wb.sheetnames:
            wb["JSR"].sheet_state = "hidden"

        # Remove existing JV sheet if present
        if new_sheet_name in wb.sheetnames:
            del wb[new_sheet_name]

        # Create new sheet and write data
        ws_jv = wb.create_sheet(title=new_sheet_name)
        for r in dataframe_to_rows(melted_df[final_cols], index=False, header=True):
            ws_jv.append(r)

        # Save to new output file
        temp_output = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(temp_output.name)
        wb.close()

        output_filename = f"{os.path.splitext(file.filename)[0]}_with_JV_{month_abbr}_{year}.xlsx"
        return FileResponse(path=temp_output.name, filename=output_filename)

    except ValueError as ve:
        raise HTTPException(status_code=400, detail=str(ve))
    finally:
        if os.path.exists(temp_input.name):
            os.remove(temp_input.name)

if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8004)